import random
from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
import openpyxl as xl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill


#Excel Sheet

wb = xl.Workbook()
ws = wb.active 
ws.title = 'Crypto Tracker'
myfont = Font(name='Times New Roman', size=14, italic=False, bold=True)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# format

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws['A1'] = 'Coin Name'
ws['B1'] = 'Symbol'
ws['C1'] = 'Price'
ws['D1'] = '24h Change'
ws['E1'] = 'Previous Price'
ws['A1'].font = myfont
ws['B1'].font = myfont
ws['C1'].font = myfont
ws['D1'].font = myfont
ws['E1'].font = myfont
ws['A1'].fill = yellow_fill
ws['B1'].fill = yellow_fill
ws['C1'].fill = yellow_fill
ws['D1'].fill = yellow_fill
ws['E1'].fill = yellow_fill

# webscraping 

url = 'https://www.livecoinwatch.com/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)
print()

crypto_data = soup.findAll('tr')


i = 2

# iterate through top 5 coins 

for row in crypto_data[2:7]:
    tag = row.findAll('td')

    # find elements needed
    
    name_ticker = tag[1].text
    raw_price = tag[2].text
    raw_change = tag[8].text

    # format needed

    name = name_ticker.split()[1]
    ticker = name_ticker.split()[0]
    price = float(raw_price.replace('$',''))
    change = float(raw_change.replace('.','').replace('%',''))
    
    # check if it is decreasing or increasing 
    # get previous price

    tag_class = tag[8].get('class')
    direction = tag_class[3]
    
    if direction == 'fall':
        real_change = (change * -1)/10000    
        previous_price = round(price * (1 + real_change),2)
        visible_change = (change/100) * (-1)
    else:
        real_change = change/10000
        previous_price = price * (1 - (real_change))
        visible_change = change/100
    
    #add elements to spreadsheet

    ws[f'A{i}'] = name
    ws[f'B{i}'] = ticker
    ws[f'C{i}'] = price
    ws[f'D{i}'] = str(visible_change) + '%'
    ws[f'D{i}'].alignment = Alignment(horizontal='right')
    ws[f'E{i}'] = previous_price
    
    i += 1 

    for cell in ws['C:C']:
        cell.number_format = '$##,##0.0000'

    for cell in ws['D:D']:
        cell.number_format = u'##0.00'

    for cell in ws['E:E']:
        cell.number_format = u'$##,##0.0000'


    wb.save('CryptoTracker.xlsx')

    import keys
    from twilio.rest import Client
    
    client = Client(keys.account_sid, keys.auth_token)
    TWnumber = "+15075981131"
    myphone = '+14076395442'
    mychoice = f'Buy {name}'
    deviation = previous_price - price
    if (name == 'Bitcoin' or name == 'Ethereum') and (deviation > 5 or deviation < -5):
        textmsg = client.messages.create(to=myphone, from_=TWnumber, body=mychoice)    
